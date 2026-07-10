from pathlib import Path
import openpyxl

path = Path(r'C:\Users\MumTy\Desktop\Inventory Tracker March.xlsx')
print('exists', path.exists())
wb = openpyxl.load_workbook(path, data_only=True)
print('SHEETS', wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n=== SHEET: {sheet_name} ===')
    for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
        if any(value is not None and str(value).strip() != '' for value in row):
            print(row)
    print('rows', ws.max_row, 'cols', ws.max_column)
